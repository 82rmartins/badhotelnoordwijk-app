#!/usr/bin/env python3
"""
Análise completa dos 4 relatórios XLSX do Bad Hotel Noordwijk
Data de referência: 25 de Fevereiro de 2026
"""

import openpyxl
from datetime import datetime, timedelta
from collections import defaultdict

TODAY = datetime(2026, 2, 25)

def analyze_daily_report():
    """Analisa o relatório DAILY para extrair ocupação por dia"""
    print("\n" + "="*80)
    print("📊 DAILY REPORT ANALYSIS")
    print("="*80)
    
    wb = openpyxl.load_workbook('/app/daily.xlsx')
    sheet = wb.active
    
    # Print headers
    print("\n📋 Headers (Row 1):")
    headers = []
    for col in range(1, 20):
        val = sheet.cell(row=1, column=col).value
        if val:
            headers.append((col, val))
            print(f"  Col {col}: {val}")
    
    # Find key columns
    date_col = None
    rooms_occupied_col = None
    rooms_available_col = None
    revenue_col = None
    
    for col, header in headers:
        header_lower = str(header).lower()
        if 'date' in header_lower or 'day' in header_lower:
            date_col = col
        if 'occupied' in header_lower or 'rooms sold' in header_lower:
            rooms_occupied_col = col
        if 'available' in header_lower:
            rooms_available_col = col
        if 'revenue' in header_lower or 'accommodation' in header_lower:
            revenue_col = col
    
    print(f"\n🔍 Detected columns:")
    print(f"  Date column: {date_col}")
    print(f"  Rooms Occupied column: {rooms_occupied_col}")
    print(f"  Rooms Available column: {rooms_available_col}")
    print(f"  Revenue column: {revenue_col}")
    
    # Extract data for next 14 days from today (Feb 25)
    print(f"\n📅 Daily data starting from {TODAY.strftime('%Y-%m-%d')}:")
    print("-" * 60)
    
    daily_data = {}
    total_rooms = None
    
    for row in range(2, sheet.max_row + 1):
        date_val = sheet.cell(row=row, column=date_col).value if date_col else None
        
        if date_val:
            # Parse date
            if isinstance(date_val, datetime):
                row_date = date_val
            elif isinstance(date_val, str):
                try:
                    row_date = datetime.strptime(date_val, '%Y-%m-%d')
                except:
                    try:
                        row_date = datetime.strptime(date_val, '%d-%m-%Y')
                    except:
                        continue
            else:
                continue
            
            # Get values
            occupied = sheet.cell(row=row, column=rooms_occupied_col).value if rooms_occupied_col else 0
            available = sheet.cell(row=row, column=rooms_available_col).value if rooms_available_col else 0
            revenue = sheet.cell(row=row, column=revenue_col).value if revenue_col else 0
            
            # Store total rooms (from available column)
            if available and total_rooms is None:
                total_rooms = available
            
            daily_data[row_date.strftime('%Y-%m-%d')] = {
                'date': row_date,
                'occupied': occupied or 0,
                'available': available or 0,
                'revenue': revenue or 0
            }
            
            # Print data for next 14 days
            if row_date >= TODAY and row_date < TODAY + timedelta(days=14):
                day_name = row_date.strftime('%A')
                print(f"  {row_date.strftime('%Y-%m-%d')} ({day_name}): {occupied} quartos ocupados, {available} disponíveis, €{revenue:.2f} receita")
    
    print(f"\n🏨 Total de quartos no hotel: {total_rooms}")
    return daily_data, total_rooms


def analyze_reservation_report():
    """Analisa o relatório de RESERVAS para extrair chegadas e partidas"""
    print("\n" + "="*80)
    print("📊 RESERVATION REPORT ANALYSIS")
    print("="*80)
    
    wb = openpyxl.load_workbook('/app/reservation.xlsx')
    sheet = wb.active
    
    # Print headers
    print("\n📋 Headers (Row 1):")
    headers = []
    for col in range(1, 30):
        val = sheet.cell(row=1, column=col).value
        if val:
            headers.append((col, val))
            print(f"  Col {col}: {val}")
    
    # Find key columns
    arrival_col = None
    departure_col = None
    status_col = None
    guest_col = None
    
    for col, header in headers:
        header_lower = str(header).lower()
        if 'arrival' in header_lower or 'check-in' in header_lower or 'chegada' in header_lower:
            arrival_col = col
        if 'departure' in header_lower or 'check-out' in header_lower or 'partida' in header_lower:
            departure_col = col
        if 'status' in header_lower or 'state' in header_lower:
            status_col = col
        if 'guest' in header_lower or 'name' in header_lower or 'customer' in header_lower:
            guest_col = col
    
    print(f"\n🔍 Detected columns:")
    print(f"  Arrival column: {arrival_col}")
    print(f"  Departure column: {departure_col}")
    print(f"  Status column: {status_col}")
    print(f"  Guest column: {guest_col}")
    
    # Extract arrivals and departures
    arrivals = defaultdict(list)
    departures = defaultdict(list)
    
    for row in range(2, sheet.max_row + 1):
        arrival_val = sheet.cell(row=row, column=arrival_col).value if arrival_col else None
        departure_val = sheet.cell(row=row, column=departure_col).value if departure_col else None
        status_val = sheet.cell(row=row, column=status_col).value if status_col else None
        guest_val = sheet.cell(row=row, column=guest_col).value if guest_col else f"Guest Row {row}"
        
        # Skip cancelled reservations
        if status_val:
            status_lower = str(status_val).lower()
            if 'cancel' in status_lower:
                continue
        
        # Parse arrival date
        if arrival_val:
            if isinstance(arrival_val, datetime):
                arr_date = arrival_val.strftime('%Y-%m-%d')
            elif isinstance(arrival_val, str):
                try:
                    arr_date = datetime.strptime(arrival_val[:10], '%Y-%m-%d').strftime('%Y-%m-%d')
                except:
                    try:
                        arr_date = datetime.strptime(arrival_val[:10], '%d-%m-%Y').strftime('%Y-%m-%d')
                    except:
                        arr_date = None
            else:
                arr_date = None
            
            if arr_date:
                arrivals[arr_date].append(guest_val)
        
        # Parse departure date
        if departure_val:
            if isinstance(departure_val, datetime):
                dep_date = departure_val.strftime('%Y-%m-%d')
            elif isinstance(departure_val, str):
                try:
                    dep_date = datetime.strptime(departure_val[:10], '%Y-%m-%d').strftime('%Y-%m-%d')
                except:
                    try:
                        dep_date = datetime.strptime(departure_val[:10], '%d-%m-%Y').strftime('%Y-%m-%d')
                    except:
                        dep_date = None
            else:
                dep_date = None
            
            if dep_date:
                departures[dep_date].append(guest_val)
    
    # Print arrivals and departures for next 7 days
    print(f"\n📥 CHEGADAS (Arrivals) próximos 7 dias:")
    print("-" * 60)
    for i in range(7):
        check_date = (TODAY + timedelta(days=i)).strftime('%Y-%m-%d')
        day_name = (TODAY + timedelta(days=i)).strftime('%A')
        count = len(arrivals.get(check_date, []))
        print(f"  {check_date} ({day_name}): {count} chegadas")
        if count > 0 and count <= 5:
            for guest in arrivals[check_date]:
                print(f"    - {guest}")
    
    print(f"\n📤 PARTIDAS (Departures) próximos 7 dias:")
    print("-" * 60)
    for i in range(7):
        check_date = (TODAY + timedelta(days=i)).strftime('%Y-%m-%d')
        day_name = (TODAY + timedelta(days=i)).strftime('%A')
        count = len(departures.get(check_date, []))
        print(f"  {check_date} ({day_name}): {count} partidas")
        if count > 0 and count <= 5:
            for guest in departures[check_date]:
                print(f"    - {guest}")
    
    return dict(arrivals), dict(departures)


def analyze_weekly_report():
    """Analisa o relatório WEEKLY"""
    print("\n" + "="*80)
    print("📊 WEEKLY REPORT ANALYSIS")
    print("="*80)
    
    wb = openpyxl.load_workbook('/app/weekly.xlsx')
    sheet = wb.active
    
    print("\n📋 Headers and first 5 rows:")
    for row in range(1, min(10, sheet.max_row + 1)):
        row_data = []
        for col in range(1, min(15, sheet.max_column + 1)):
            val = sheet.cell(row=row, column=col).value
            if val is not None:
                row_data.append(f"[{col}]{val}")
        if row_data:
            print(f"  Row {row}: {', '.join(row_data)}")


def analyze_monthly_report():
    """Analisa o relatório MONTHLY"""
    print("\n" + "="*80)
    print("📊 MONTHLY REPORT ANALYSIS")
    print("="*80)
    
    wb = openpyxl.load_workbook('/app/monthly.xlsx')
    sheet = wb.active
    
    print("\n📋 Headers and first 5 rows:")
    for row in range(1, min(10, sheet.max_row + 1)):
        row_data = []
        for col in range(1, min(15, sheet.max_column + 1)):
            val = sheet.cell(row=row, column=col).value
            if val is not None:
                row_data.append(f"[{col}]{val}")
        if row_data:
            print(f"  Row {row}: {', '.join(row_data)}")


def main():
    print("\n" + "🏨 "*20)
    print("    BAD HOTEL NOORDWIJK - ANÁLISE COMPLETA DOS RELATÓRIOS")
    print("    Data de Referência: 25 de Fevereiro de 2026")
    print("🏨 "*20)
    
    # Analyze all reports
    daily_data, total_rooms = analyze_daily_report()
    arrivals, departures = analyze_reservation_report()
    analyze_weekly_report()
    analyze_monthly_report()
    
    # Summary
    print("\n" + "="*80)
    print("📝 RESUMO - VALORES CORRETOS PARA O DASHBOARD")
    print("="*80)
    
    print(f"\n🏨 TOTAL DE QUARTOS: {total_rooms}")
    
    print(f"\n📅 HOJE ({TODAY.strftime('%Y-%m-%d')}, Quarta-feira):")
    today_key = TODAY.strftime('%Y-%m-%d')
    if today_key in daily_data:
        d = daily_data[today_key]
        print(f"  - Quartos Ocupados: {d['occupied']}")
        print(f"  - Quartos Disponíveis: {d['available']}")
        print(f"  - Receita: €{d['revenue']:.2f}")
    print(f"  - Chegadas: {len(arrivals.get(today_key, []))}")
    print(f"  - Partidas: {len(departures.get(today_key, []))}")
    
    print(f"\n📅 AMANHÃ ({(TODAY + timedelta(days=1)).strftime('%Y-%m-%d')}, Quinta-feira):")
    tomorrow_key = (TODAY + timedelta(days=1)).strftime('%Y-%m-%d')
    if tomorrow_key in daily_data:
        d = daily_data[tomorrow_key]
        print(f"  - Quartos Ocupados: {d['occupied']}")
        print(f"  - Receita: €{d['revenue']:.2f}")
    print(f"  - Chegadas: {len(arrivals.get(tomorrow_key, []))}")
    print(f"  - Partidas: {len(departures.get(tomorrow_key, []))}")
    
    # Week data
    print(f"\n📆 RADAR - PRÓXIMOS 14 DIAS:")
    for i in range(14):
        check_date = TODAY + timedelta(days=i)
        check_key = check_date.strftime('%Y-%m-%d')
        day_name = check_date.strftime('%a')
        
        occupied = 0
        if check_key in daily_data:
            occupied = daily_data[check_key]['occupied']
        
        arr_count = len(arrivals.get(check_key, []))
        dep_count = len(departures.get(check_key, []))
        
        print(f"  {check_key} ({day_name}): {occupied} quartos | {arr_count} chegadas | {dep_count} partidas")


if __name__ == '__main__':
    main()
